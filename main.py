"""
This script creates clusters of SNOMED codes based on the transitive closure table in the SNOMED database.
The clusters are created by specifying a list of parent codes and then finding all the child codes of those parents.
The child codes are then converted to their current code if they have been retired.
The clusters are output to a csv file and an xlsx file.
The csv file is then used to create a txt file for each cluster.

Instructions:
Obtain the SNOMED databases from the NHS Digital TRUD website.
Configure the paths to the databases and adjust output file paths if necessary.
Configure the clusters dictionary with the clusters to create.
"""

import warnings
import os
import datetime
import logging
import pandas as pd
import pyodbc
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# Suppress all UserWarning category warning to prevent pandas spam re; SQLAlchemy
warnings.simplefilter('ignore', category=UserWarning)

# Setup logging and timer
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
start_time = datetime.datetime.now()

# Configuration
transitive_closure_db = r"C:\Users\eddie\NHS\HealtheAnalytics Workstream - LTC LCS Workstream\Product Specifications\Data modelling\DMWB NHS SNOMED Transitive Closure.mdb"
history_db = r"C:\Users\eddie\NHS\HealtheAnalytics Workstream - LTC LCS Workstream\Product Specifications\Data modelling\DMWB NHS SNOMED History.mdb"
snomed_db = r"C:\Users\eddie\NHS\HealtheAnalytics Workstream - LTC LCS Workstream\Product Specifications\Data modelling\DMWB NHS SNOMED.mdb"
csv_filename = 'snomed_hierarchical_clusters.csv'
xlsx_filename = 'snomed_hierarchical_clusters.xlsx'
clusters_dir = 'clusters'

# Dictionary of clusters to create
clusters = {
    'dm_cod': [73211009],
    'fh_cvd_cod': [266894000],
    'pain_cod': [276435006],
    'msk_cod': [106028002, 301366005, 421060004, 72696002, 106030000, 302258001, 302293008, 298339004, 298325004, 298343000],
}

def connect_to_db(db_path):
    conn_str = (
        r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};'
        r'DBQ=' + db_path + ';'
    )
    return pyodbc.connect(conn_str)

def get_child_codes(parent_code, conn):
    query = f"SELECT SubtypeID FROM SCTTC WHERE SuperTypeID = '{parent_code}'"
    return pd.read_sql(query, conn)

def get_current_code(old_code, conn):
    query = f"SELECT NEWCUI FROM SCTHIST WHERE OLDCUI = '{old_code}'"
    return pd.read_sql(query, conn)

def get_term_for_code(code, conn):
    query = f"SELECT TERM FROM SCT WHERE CUI = '{code}'"
    return pd.read_sql(query, conn)

def chunk_list(lst, size):
    for i in range(0, len(lst), size):
        yield lst[i:i + size]

def fetch_all_child_codes(parent_codes, conn_transitive, conn_history, conn_snomed, exceptions=None):
    if exceptions is None:
        exceptions = set()

    all_codes = set(parent_codes)
    newly_added = set(parent_codes)

    while newly_added:
        temp_new_children = set()
        for chunk in chunk_list(list(newly_added), 500):
            query = "SELECT SubtypeID FROM SCTTC WHERE SuperTypeID IN ({})".format(",".join(['?'] * len(chunk)))
            cursor = conn_transitive.cursor()
            cursor.execute(query, chunk)
            temp_new_children.update({row.SubtypeID for row in cursor.fetchall()})

        newly_added = temp_new_children - all_codes
        newly_added -= exceptions
        all_codes.update(newly_added)

    # Convert old codes to current codes and filter duplicates
    final_codes = set()
    for code in all_codes:
        current_code = get_current_code(code, conn_history)
        if not current_code.empty:
            code = current_code.iloc[0]['NEWCUI']

        final_codes.add(code)

    return final_codes

def output_to_csv(cluster_dict, filename):
    conn_transitive = connect_to_db(transitive_closure_db)
    conn_history = connect_to_db(history_db)
    conn_snomed = connect_to_db(snomed_db)

    all_child_codes = []
    total_codes_count = 0

    for cluster_id, parent_codes in cluster_dict.items():
        child_codes = fetch_all_child_codes(parent_codes, conn_transitive, conn_history, conn_snomed)
        codes_count = len(child_codes)
        total_codes_count += codes_count
        for code in child_codes:
            term = get_term_for_code(code, conn_snomed)
            term_value = term.iloc[0]['TERM'] if not term.empty else 'Unknown'
            all_child_codes.append((cluster_id, str(code), term_value))
        logging.info(f"Cluster '{cluster_id}' created with {codes_count} codes.")

    df = pd.DataFrame(all_child_codes, columns=['Cluster ID', 'Concept ID', 'Term'])

    df.to_csv(filename, index=False, encoding='utf-8-sig')

    conn_transitive.close()
    conn_history.close()
    conn_snomed.close()

    logging.info (f"Completed creating clusters with {total_codes_count} total codes across {len(cluster_dict)} clusters.")
    logging.info(f"Csv file '{filename}' created.")


def create_xlsx_from_csv(csv_file, xlsx_file):
    df = pd.read_csv(csv_file, dtype={'Concept ID': str})

    with pd.ExcelWriter(xlsx_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Clusters')

        workbook = writer.book
        worksheet = writer.sheets['Clusters']

        # Set Concept ID format to text
        for row in worksheet.iter_rows(min_row=2, max_col=2, max_row=worksheet.max_row):
            for cell in row:
                cell.number_format = '@'

        # Create a table over the data range
        max_column = get_column_letter(worksheet.max_column)
        data_range = f"A1:{max_column}{worksheet.max_row}"
        table = Table(displayName="ClusterTable", ref=data_range)

        # Add default table style
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        worksheet.add_table(table)

    logging.info(f"Xlsx file '{xlsx_file}' created.")

def create_txt_files_from_csv(csv_filename, output_dir):
    # Read the CSV file
    df = pd.read_csv(csv_filename)

    # Create the output directory if it doesn't exist
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # Group by 'Cluster ID' and save each group to a .txt file
    for cluster_id, group in df.groupby('Cluster ID'):
        txt_filename = os.path.join(output_dir, f"{cluster_id}.txt")
        with open(txt_filename, 'w') as f:
            formatted_codes = ','.join([f"'{code}'" for code in group['Concept ID']])
            f.write(formatted_codes)
        logging.info(f"Txt file '{txt_filename}' created.")

# Main script execution
logging.info("Script started. Found {}".format(len(clusters)) + " clusters to create.")
output_to_csv(clusters, csv_filename)
create_xlsx_from_csv(csv_filename, xlsx_filename)
create_txt_files_from_csv(csv_filename, clusters_dir)

end_time = datetime.datetime.now()
execution_time = end_time - start_time
total_seconds = execution_time.total_seconds()
minutes = int(total_seconds // 60)
seconds = int(total_seconds % 60) 

if minutes > 0:
    time_str = f"{minutes} minutes and {seconds} seconds"
else:
    time_str = f"{seconds} seconds"


logging.info(f"Script executed in {time_str}")
